from __future__ import annotations

import csv
import json
import re
import zipfile
from collections import Counter
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape

import pandas as pd
from docx import Document
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from pypdf import PdfReader
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, TableStyle


ROOT = Path(__file__).resolve().parents[1]
TMP_DIR = ROOT / "tmp" / "hmrc_tdr_audit"
TECH_DOC_ROOT = TMP_DIR / "cds_docs" / "CDS_Technical_Documentation_20260320"
OUTPUT_DIR = ROOT / "documentation" / "hmrc_tdr_audit"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TECH_DOCS_ZIP_URL = (
    "https://developer.service.hmrc.gov.uk/guides/customs-declarations-end-to-end-service-guide/"
    "documentation/resources/CDS_Technical_Documentation.zip"
)


@dataclass
class Source:
    source_id: str
    title: str
    version: str
    url: str
    last_updated: str
    publication_type: str
    identifier_hint: str
    local_path: str | None = None
    notes: str | None = None


@dataclass
class Rule:
    sequential_id: int
    traceability_id: str
    category: str
    scope: str
    description: str
    hmrc_reference: str
    source_id: str
    source_title: str
    source_version: str
    page_or_section: str
    url: str
    last_updated: str
    error_code: str
    acceptance_criteria: list[str]
    verification_steps: list[str]
    pass_fail_status: str = "Not-tested"
    evidence_link: str = ""
    extra: dict[str, str] = field(default_factory=dict)


def normalise_ws(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\s*\n\s*", "\n", text)
    return text.strip()


def compact(value: object) -> str:
    return normalise_ws(value).replace("\n", " ")


def docx_paragraphs(path: Path) -> list[str]:
    doc = Document(path)
    return [compact(p.text) for p in doc.paragraphs if compact(p.text)]


def odf_text(path: Path) -> list[str]:
    texts: list[str] = []
    with zipfile.ZipFile(path) as zf:
        content = ET.fromstring(zf.read("content.xml"))
        for elem in content.iter():
            if elem.text and elem.text.strip():
                texts.append(compact(elem.text))
    return texts


def pptx_slides(path: Path) -> list[str]:
    slides: list[str] = []
    with zipfile.ZipFile(path) as zf:
        names = sorted(
            n for n in zf.namelist() if n.startswith("ppt/slides/slide") and n.endswith(".xml")
        )
        for name in names:
            root = ET.fromstring(zf.read(name))
            parts = []
            for elem in root.iter():
                if elem.text and elem.text.strip():
                    parts.append(compact(elem.text))
            slides.append(" | ".join(parts))
    return slides


def pdf_text(path: Path) -> list[str]:
    reader = PdfReader(str(path))
    lines: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        for line in text.splitlines():
            line = compact(line)
            if line:
                lines.append(line)
    return lines


def build_sources() -> dict[str, Source]:
    sources: dict[str, Source] = {}

    def add(source: Source) -> None:
        sources[source.source_id] = source

    add(
        Source(
            source_id="web_dev_setup",
            title="Developer set up | Customs Declarations End-to-End Service Guide",
            version="Web page",
            url="https://developer.service.hmrc.gov.uk/guides/customs-declarations-end-to-end-service-guide/documentation/set-up-developers.html",
            last_updated="2026-03-20",
            publication_type="Developer guide",
            identifier_hint="HTML section anchors",
            notes="Contains Trade Test, Trader Dress Rehearsal, API versioning, rate limiting, and error handling sections.",
        )
    )
    add(
        Source(
            source_id="web_td_rehearsal",
            title="Using the Trader Dress Rehearsal service",
            version="Web page",
            url="https://www.gov.uk/guidance/using-the-trader-dress-rehearsal-service",
            last_updated="2023-09-15",
            publication_type="GOV.UK guidance",
            identifier_hint="Section heading or update item",
        )
    )
    add(
        Source(
            source_id="web_service_availability",
            title="Customs Declaration Service: service availability and issues",
            version="Web page",
            url="https://www.gov.uk/guidance/customs-declaration-service-service-availability-and-issues",
            last_updated="2026-03-25",
            publication_type="GOV.UK guidance",
            identifier_hint="Update entry date",
        )
    )
    add(
        Source(
            source_id="web_report_problem",
            title="Report a problem using the Customs Declaration Service",
            version="Web page",
            url="https://www.gov.uk/guidance/report-a-problem-using-the-customs-declaration-service",
            last_updated="2026-03-25",
            publication_type="GOV.UK guidance",
            identifier_hint="Section heading or update item",
        )
    )
    add(
        Source(
            source_id="path_to_production",
            title="The CDS Path to Production",
            version="February 2024",
            url="https://developer.service.hmrc.gov.uk/guides/customs-declarations-end-to-end-service-guide/documentation/resources/The_Path_to_Production_2024.pdf",
            last_updated="2024-02-01",
            publication_type="PDF guidance",
            identifier_hint="PDF page number",
            local_path=str(TMP_DIR / "The_Path_to_Production_2024.pdf"),
        )
    )
    add(
        Source(
            source_id="tech_change_log",
            title="CDS Technical Documentation - Change Log",
            version="20260319",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2026-03-20",
            publication_type="Bundle change log",
            identifier_hint="Bundle path + issue date entry",
            local_path=str(TECH_DOC_ROOT / "CDS Technical Documentation - Change Log.20260319.odt"),
        )
    )
    add(
        Source(
            source_id="dssd",
            title="CDS Declaration Submission Service Design",
            version="v2.30",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2025-12-18",
            publication_type="Service design specification",
            identifier_hint="Bundle path + paragraph heading",
            local_path=str(TECH_DOC_ROOT / "CDS 03 - Customs Declaration" / "CDS 03 DSSD v2.30.docx"),
        )
    )
    add(
        Source(
            source_id="imports_tcm",
            title="CDS 03 Declaration Technical Completion Matrix for Imports",
            version="v3.92",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2026-02-19",
            publication_type="Validation matrix",
            identifier_hint="Workbook sheet + DE or cell reference",
            local_path=str(
                TECH_DOC_ROOT / "CDS 03 - Customs Declaration" / "CDS 03 Declaration TCM for Imports v3.92.xlsx"
            ),
        )
    )
    add(
        Source(
            source_id="exports_tcm",
            title="CDS 03 Declaration Technical Completion Matrix for Exports",
            version="v5.7",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2025-05-01",
            publication_type="Validation matrix",
            identifier_hint="Workbook sheet + DE or cell reference",
            local_path=str(
                TECH_DOC_ROOT / "CDS 03 - Customs Declaration" / "CDS 03 Declaration TCM for Exports v5.7.xlsx"
            ),
        )
    )
    add(
        Source(
            source_id="additional_message_tcm",
            title="CDS Additional Message Technical Completion Matrix",
            version="v1.4",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2025-09-17",
            publication_type="Validation matrix",
            identifier_hint="Workbook sheet + XML path",
            local_path=str(
                TECH_DOC_ROOT / "CDS 03 - Customs Declaration" / "CDS Additional Message Technical Completion Matrix v1.4.ods"
            ),
        )
    )
    add(
        Source(
            source_id="codelists",
            title="CDS 03 CDS Codelists and WCO References 5.1.0",
            version="v2.55",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2026-03-19",
            publication_type="Codelist workbook",
            identifier_hint="Workbook sheet + codelist name or WCO path",
            local_path=str(
                TECH_DOC_ROOT / "CDS 03 - Customs Declaration" / "CDS 03 CDS Codelists and WCO References 5.1.0 v2.55.xlsx"
            ),
        )
    )
    add(
        Source(
            source_id="dis",
            title="CDS Declaration Information Service",
            version="v2.97",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2025-06-12",
            publication_type="Service design specification",
            identifier_hint="Bundle path + paragraph heading",
            local_path=str(
                TECH_DOC_ROOT / "CDS 03 - Customs Declaration" / "CDS Declaration Information Service v2.97.docx"
            ),
        )
    )
    add(
        Source(
            source_id="validation_authorisation",
            title="CDS Guidance for Software Developers - Validation of Authorisations",
            version="v1.9",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2020-07-22",
            publication_type="Guidance note",
            identifier_hint="Bundle path + authorisation type",
            local_path=str(
                TECH_DOC_ROOT / "CDS 11 - Guidance Documents" / "CDS Guidance for Developers - Validation for Authorisation v1.9.odt"
            ),
        )
    )
    add(
        Source(
            source_id="mucr_format",
            title="CDS Guidance for Software Developers - MUCR Format",
            version="Baselined",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2023-04-20",
            publication_type="Guidance note",
            identifier_hint="Bundle path + section",
            local_path=str(
                TECH_DOC_ROOT / "CDS Supporting Documentation" / "CDS Guidance for Developers - MUCR format.docx"
            ),
        )
    )
    add(
        Source(
            source_id="rounding",
            title="Tax Calculation Processing - Rounding in CDS",
            version="v1.5",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2023-05-11",
            publication_type="Guidance note",
            identifier_hint="Bundle path + section",
            local_path=str(
                TECH_DOC_ROOT / "CDS Supporting Documentation" / "Tax Calculation Processing - Rounding in CDS v1.5.docx"
            ),
        )
    )
    add(
        Source(
            source_id="waf_examples",
            title="WAF Trigger Examples",
            version="v0.4",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2025-01-08",
            publication_type="Validation bulletin",
            identifier_hint="Workbook sheet + rule/example ID",
            local_path=str(TECH_DOC_ROOT / "CDS Supporting Documentation" / "WAF Trigger Examples v0.4.xlsx"),
        )
    )
    add(
        Source(
            source_id="file_upload_ui",
            title="CDS Secure File Upload UI User Guide",
            version="v0.44",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2023-01-05",
            publication_type="Guidance note",
            identifier_hint="Bundle path + section",
            local_path=str(
                TECH_DOC_ROOT / "CDS Supporting Documentation" / "CDS Secure File Upload UI User Guide v0.44.odt"
            ),
        )
    )
    add(
        Source(
            source_id="runbook",
            title="CDS RunBook",
            version="v4.3",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2024-02-07",
            publication_type="Developer bulletin",
            identifier_hint="Slide number",
            local_path=str(TECH_DOC_ROOT / "CDS Supporting Documentation" / "CDS RunBook V4.3.pptx"),
        )
    )
    add(
        Source(
            source_id="tt_tdr_scope",
            title="CDS External TT TDR Scope Guidance",
            version="2025-12-04 issue",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2025-12-02",
            publication_type="Scope bulletin",
            identifier_hint="Slide number",
            local_path=str(TECH_DOC_ROOT / "CDS Supporting Documentation" / "CDS External TT TDR Scope Guidance.pptx"),
        )
    )
    add(
        Source(
            source_id="ttm33_scope",
            title="CDS TTM33.0 Scope Guidance",
            version="2026-03-20 issue",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2026-03-19",
            publication_type="Scope bulletin",
            identifier_hint="Slide number",
            local_path=str(TECH_DOC_ROOT / "CDS Supporting Documentation" / "CDS TTM33.0 Scope Guidance 20260320.pptx"),
        )
    )
    add(
        Source(
            source_id="test_data_library",
            title="CDS Test Data Library",
            version="2025-10-07 issue",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2025-10-07",
            publication_type="Test data workbook",
            identifier_hint="Workbook sheet + trader or note",
            local_path=str(TECH_DOC_ROOT / "CDS Supporting Documentation" / "CDS Test Data Library_20251007.xlsx"),
        )
    )
    add(
        Source(
            source_id="trade_test_kel",
            title="CDS Trade Test Known Error Log",
            version="2026-03-19 issue",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2026-03-19",
            publication_type="Known error log",
            identifier_hint="Workbook sheet + KEL ID",
            local_path=str(TECH_DOC_ROOT / "CDS Supporting Documentation" / "CDS_Trade_Test_KEL_20260319.xlsm"),
        )
    )
    add(
        Source(
            source_id="trade_test_query_log",
            title="CDS Trade Test Query Log",
            version="v5.38",
            url=TECH_DOCS_ZIP_URL,
            last_updated="2025-05-13",
            publication_type="Developer bulletin",
            identifier_hint="Workbook sheet + log entry",
            local_path=str(TECH_DOC_ROOT / "CDS Supporting Documentation" / "CDS Trade Test Query Log V5.38.xlsm"),
        )
    )
    add(
        Source(
            source_id="error_codes",
            title="CDS Error Codes",
            version="11-03-2026 issue",
            url="https://assets.publishing.service.gov.uk/media/69b41375b84f01b2be53a207/CDS_Error_Codes_11-03-2026.ods",
            last_updated="2026-03-11",
            publication_type="Error code list",
            identifier_hint="ODS row by code",
            local_path=str(TMP_DIR / "CDS_Error_Codes_11-03-2026.ods"),
        )
    )
    add(
        Source(
            source_id="known_differences",
            title="Declarant Known Differences and Workarounds Pre 5.0.0",
            version="Pre 5.0.0 / 24-11-2025 update",
            url="https://assets.publishing.service.gov.uk/media/6929d428a245b0985f0342ef/Declarant_CWD_Pre_5.0.0.ods",
            last_updated="2025-11-24",
            publication_type="Known difference log",
            identifier_hint="ODS tab + KD/KEL entry",
            local_path=str(TMP_DIR / "Declarant_CWD_Pre_5.0.0.ods"),
        )
    )
    return sources


def extract_code_list_name(description: str) -> str:
    match = re.search(r"CodeList ([A-Za-z0-9_()\- ]+)", description)
    if match:
        return compact(match.group(1))
    match = re.search(r"codelist ([A-Za-z0-9_#()\- ]+)", description, re.I)
    if match:
        return compact(match.group(1))
    return ""


OBLIGATION_LABELS = {
    "M": "mandatory",
    "M*": "mandatory when the group-level item is present",
    "M**": "mandatory for green-lane movements",
    "D": "dependent",
    "D*": "dependent for green-lane movements",
    "O": "optional",
    "NA": "not allowed",
    "N/A": "not applicable",
}


def field_error_pattern(obligation: str, format_value: str, code_list_name: str, category: str) -> str:
    if category == "waf":
        return "HTTP 403 PAYLOAD_FORBIDDEN | HTTP 500 INTERNAL_SERVER_ERROR"
    if category == "schema":
        return "BAD_REQUEST | xml_validation_error"
    codes: list[str] = []
    obligation = compact(obligation)
    if obligation in {"M", "M*", "M**"}:
        codes.append("CDS10001")
    if obligation in {"D", "D*"}:
        codes.append("CDS12070")
    if obligation == "NA":
        codes.append("CDS10002")
    if format_value and format_value not in {"N/A", "NA"}:
        codes.extend(["CDS10010", "xml_validation_error"])
    if code_list_name:
        codes.append("CDS10020")
    return " | ".join(dict.fromkeys(codes)) if codes else "DMSREJ validationResultType (rule-dependent)"


def acceptance_for_field(
    scope: str,
    declaration_kind: str,
    data_element: str,
    field_name: str,
    xml_path: str,
    level: str,
    cardinality: str,
    format_value: str,
    code_list_name: str,
    obligation: str,
) -> list[str]:
    obligation_text = OBLIGATION_LABELS.get(obligation, compact(obligation).lower() or "rule-dependent")
    criteria = [
        f"{scope} context is {declaration_kind}.",
        f"{data_element} {field_name} at {level} level is {obligation_text}.",
    ]
    if xml_path:
        criteria.append(f"Populate XML path {xml_path} only where this rule permits it.")
    if cardinality:
        criteria.append(f"Respect cardinality {cardinality}.")
    if format_value and format_value not in {"N/A", "NA"}:
        criteria.append(f"Value conforms to HMRC format {format_value}.")
    if code_list_name:
        criteria.append(f"Value is drawn from HMRC codelist {code_list_name}.")
    return criteria


def verification_steps_for_field(scope: str, declaration_kind: str, data_element: str, field_name: str) -> list[str]:
    return [
        f"Create a representative {scope.lower()} payload for {declaration_kind}.",
        f"Populate or omit {data_element} {field_name} according to the rule under test.",
        "Run local schema validation against the HMRC-published XSD where applicable.",
        "Run business-rule validation or submit in the correct HMRC environment for the message type.",
        "Record whether the payload is accepted without the listed error code or pattern.",
    ]


def looks_like_data_element(value: object) -> bool:
    text = compact(value)
    return bool(re.fullmatch(r"\d+/\d+[A-Z]?", text))


def build_error_code_map(path: Path) -> dict[str, dict[str, str]]:
    xl = pd.ExcelFile(path, engine="odf")
    df = xl.parse(xl.sheet_names[0])
    error_map: dict[str, dict[str, str]] = {}
    for _, row in df.iterrows():
        code = compact(row.iloc[0] if len(row) > 0 else "")
        if not code.startswith("CDS"):
            continue
        error_map[code] = {
            "description": compact(row.iloc[1] if len(row) > 1 else ""),
            "explanation": compact(row.iloc[2] if len(row) > 2 else ""),
        }
    return error_map


def parse_import_field_rules(source: Source, next_id: int) -> tuple[list[dict[str, str]], list[Rule], int]:
    workbook = load_workbook(source.local_path, read_only=True, data_only=True)
    ws = workbook["Data Completion Rules"]

    header_row_idx = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [compact(v) for v in row]
        if "UCC Data Element Order No." in values:
            header_row_idx = idx
            break
    if header_row_idx is None:
        raise RuntimeError("Imports TCM header row not found")

    proc_row_idx = header_row_idx + 1
    proc_row = [compact(ws.cell(proc_row_idx, c).value) for c in range(1, ws.max_column + 1)]
    categories = [(c, proc_row[c - 1]) for c in range(11, ws.max_column + 1) if proc_row[c - 1]]

    field_catalog: list[dict[str, str]] = []
    rules: list[Rule] = []

    for row_idx in range(proc_row_idx + 1, ws.max_row + 1):
        row = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        if not looks_like_data_element(row[1]):
            continue
        data_element = compact(row[1])
        field_name = compact(row[2])
        xml_path = compact(row[3])
        level = compact(row[5])
        cardinality = compact(row[6])
        format_value = compact(row[7])
        description = compact(row[9])
        code_list_name = extract_code_list_name(description)
        field_entry = {
            "direction": "Import",
            "de": data_element,
            "field_name": field_name,
            "xml_path": xml_path,
            "level": level,
            "cardinality": cardinality,
            "format": format_value,
            "code_list": code_list_name,
            "source_ref": f"{source.title} | {source.version} | Data Completion Rules!{row_idx}",
        }
        for col_idx, proc_name in categories:
            obligation = compact(row[col_idx - 1])
            if obligation:
                field_entry[f"obligation_{proc_name}"] = obligation
        field_catalog.append(field_entry)

        for col_idx, proc_name in categories:
            obligation = compact(row[col_idx - 1])
            if not obligation or obligation == "N/A":
                continue
            desc = (
                f"Import {proc_name}: DE {data_element} {field_name} at {level} level is "
                f"{OBLIGATION_LABELS.get(obligation, obligation)}; cardinality {cardinality or 'n/a'}; "
                f"format {format_value or 'n/a'}."
            )[:255]
            rules.append(
                Rule(
                    sequential_id=next_id,
                    traceability_id=f"TDR-R-{next_id:05d}",
                    category="Field-level requirement",
                    scope="Import declaration",
                    description=desc,
                    hmrc_reference=f"{source.title} | {source.version} | Data Completion Rules!{row_idx}",
                    source_id=source.source_id,
                    source_title=source.title,
                    source_version=source.version,
                    page_or_section=f"Data Completion Rules row {row_idx}",
                    url=source.url,
                    last_updated=source.last_updated,
                    error_code=field_error_pattern(obligation, format_value, code_list_name, "field"),
                    acceptance_criteria=acceptance_for_field(
                        "Import declaration",
                        proc_name,
                        data_element,
                        field_name,
                        xml_path,
                        level,
                        cardinality,
                        format_value,
                        code_list_name,
                        obligation,
                    ),
                    verification_steps=verification_steps_for_field(
                        "Import declaration", proc_name, data_element, field_name
                    ),
                    extra={
                        "obligation": obligation,
                        "xml_path": xml_path,
                        "format": format_value,
                        "code_list": code_list_name,
                    },
                )
            )
            next_id += 1

    return field_catalog, rules, next_id


def parse_export_field_rules(source: Source, next_id: int) -> tuple[list[dict[str, str]], list[Rule], int]:
    workbook = load_workbook(source.local_path, read_only=True, data_only=True)
    ws = workbook["Data Completion Rules"]

    header_row_idx = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [compact(v) for v in row]
        if "UCC Data Element Order No." in values:
            header_row_idx = idx
            break
    if header_row_idx is None:
        raise RuntimeError("Exports TCM header row not found")

    header_row = [compact(ws.cell(header_row_idx, c).value) for c in range(1, ws.max_column + 1)]
    categories = [(c, header_row[c - 1]) for c in range(12, ws.max_column + 1) if header_row[c - 1]]

    field_catalog: list[dict[str, str]] = []
    rules: list[Rule] = []

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        if not looks_like_data_element(row[2]):
            continue
        data_element = compact(row[2])
        field_name = compact(row[3])
        xml_path = compact(row[4])
        level = compact(row[6])
        cardinality = compact(row[7])
        format_value = compact(row[8])
        description = compact(row[10])
        code_list_name = extract_code_list_name(description)
        field_entry = {
            "direction": "Export",
            "de": data_element,
            "field_name": field_name,
            "xml_path": xml_path,
            "level": level,
            "cardinality": cardinality,
            "format": format_value,
            "code_list": code_list_name,
            "source_ref": f"{source.title} | {source.version} | Data Completion Rules!{row_idx}",
        }
        for col_idx, proc_name in categories:
            obligation = compact(row[col_idx - 1])
            if obligation:
                field_entry[f"obligation_{proc_name}"] = obligation
        field_catalog.append(field_entry)

        for col_idx, proc_name in categories:
            obligation = compact(row[col_idx - 1])
            if not obligation or obligation == "N/A":
                continue
            desc = (
                f"Export {proc_name}: DE {data_element} {field_name} at {level} level is "
                f"{OBLIGATION_LABELS.get(obligation, obligation)}; cardinality {cardinality or 'n/a'}; "
                f"format {format_value or 'n/a'}."
            )[:255]
            rules.append(
                Rule(
                    sequential_id=next_id,
                    traceability_id=f"TDR-R-{next_id:05d}",
                    category="Field-level requirement",
                    scope="Export declaration",
                    description=desc,
                    hmrc_reference=f"{source.title} | {source.version} | Data Completion Rules!{row_idx}",
                    source_id=source.source_id,
                    source_title=source.title,
                    source_version=source.version,
                    page_or_section=f"Data Completion Rules row {row_idx}",
                    url=source.url,
                    last_updated=source.last_updated,
                    error_code=field_error_pattern(obligation, format_value, code_list_name, "field"),
                    acceptance_criteria=acceptance_for_field(
                        "Export declaration",
                        proc_name,
                        data_element,
                        field_name,
                        xml_path,
                        level,
                        cardinality,
                        format_value,
                        code_list_name,
                        obligation,
                    ),
                    verification_steps=verification_steps_for_field(
                        "Export declaration", proc_name, data_element, field_name
                    ),
                    extra={
                        "obligation": obligation,
                        "xml_path": xml_path,
                        "format": format_value,
                        "code_list": code_list_name,
                    },
                )
            )
            next_id += 1

    return field_catalog, rules, next_id


def parse_additional_message_rules(source: Source, next_id: int) -> tuple[list[dict[str, str]], list[Rule], int]:
    xl = pd.ExcelFile(source.local_path, engine="odf")
    df = xl.parse("Data_Completion_Rules", header=None)
    message_types = [compact(v) for v in df.iloc[1, 4:7].tolist()]

    field_catalog: list[dict[str, str]] = []
    rules: list[Rule] = []

    for idx in range(2, len(df)):
        xml_path = compact(df.iloc[idx, 0])
        if not xml_path.startswith("Declaration/"):
            continue
        format_value = compact(df.iloc[idx, 1])
        code_list_name = compact(df.iloc[idx, 2]) if compact(df.iloc[idx, 2]) != "N" else ""
        description = compact(df.iloc[idx, 3])
        if not code_list_name:
            code_list_name = extract_code_list_name(description)
        field_entry = {
            "direction": "Additional message",
            "de": "",
            "field_name": xml_path.split("/")[-1],
            "xml_path": xml_path,
            "level": "Message",
            "cardinality": "",
            "format": format_value,
            "code_list": code_list_name,
            "source_ref": f"{source.title} | {source.version} | Data_Completion_Rules!{idx + 1}",
        }
        for offset, message_type in enumerate(message_types, start=4):
            obligation = compact(df.iloc[idx, offset])
            if obligation:
                field_entry[f"obligation_{message_type}"] = obligation
        field_catalog.append(field_entry)

        for offset, message_type in enumerate(message_types, start=4):
            obligation = compact(df.iloc[idx, offset])
            if not obligation or obligation == "N/A":
                continue
            field_name = xml_path.split("/")[-1]
            desc = (
                f"{message_type}: {field_name} ({xml_path}) is {OBLIGATION_LABELS.get(obligation, obligation)}; "
                f"format {format_value or 'n/a'}."
            )[:255]
            rules.append(
                Rule(
                    sequential_id=next_id,
                    traceability_id=f"TDR-R-{next_id:05d}",
                    category="Additional-message requirement",
                    scope="Additional message",
                    description=desc,
                    hmrc_reference=f"{source.title} | {source.version} | Data_Completion_Rules!{idx + 1}",
                    source_id=source.source_id,
                    source_title=source.title,
                    source_version=source.version,
                    page_or_section=f"Data_Completion_Rules row {idx + 1}",
                    url=source.url,
                    last_updated=source.last_updated,
                    error_code=field_error_pattern(obligation, format_value, code_list_name, "field"),
                    acceptance_criteria=acceptance_for_field(
                        "Additional message",
                        message_type,
                        xml_path,
                        field_name,
                        xml_path,
                        "Message",
                        "",
                        format_value,
                        code_list_name,
                        obligation,
                    ),
                    verification_steps=verification_steps_for_field(
                        "Additional message", message_type, xml_path, field_name
                    ),
                    extra={"obligation": obligation, "xml_path": xml_path, "format": format_value},
                )
            )
            next_id += 1

    return field_catalog, rules, next_id


def parse_import_derivation_rules(source: Source, next_id: int) -> tuple[list[Rule], int]:
    workbook = load_workbook(source.local_path, read_only=True, data_only=True)
    ws = workbook["Procedure Category Derivation 2"]

    procedure_codes = [compact(ws.cell(3, c).value) for c in range(1, ws.max_column + 1)]
    rules: list[Rule] = []
    current_decl_type = ""

    for row_idx in range(4, ws.max_row + 1):
        decl_type = compact(ws.cell(row_idx, 1).value)
        add_type = compact(ws.cell(row_idx, 2).value)
        if decl_type:
            current_decl_type = decl_type
        if not current_decl_type or not add_type:
            continue
        for col_idx in range(3, ws.max_column + 1):
            proc_code = procedure_codes[col_idx - 1]
            category = compact(ws.cell(row_idx, col_idx).value)
            if not proc_code or not category:
                continue
            desc = (
                f"Import procedure derivation: declaration type {current_decl_type}, additional declaration type "
                f"{add_type}, and requested procedure {proc_code} must derive category {category}."
            )[:255]
            rules.append(
                Rule(
                    sequential_id=next_id,
                    traceability_id=f"TDR-R-{next_id:05d}",
                    category="Cross-field integrity",
                    scope="Import declaration",
                    description=desc,
                    hmrc_reference=f"{source.title} | {source.version} | Procedure Category Derivation 2!{row_idx}",
                    source_id=source.source_id,
                    source_title=source.title,
                    source_version=source.version,
                    page_or_section=f"Procedure Category Derivation 2 row {row_idx}",
                    url=source.url,
                    last_updated=source.last_updated,
                    error_code="CDS11004 | CDS11005 | CDS12050 | CDS12052",
                    acceptance_criteria=[
                        f"DE 1/1 equals {current_decl_type}.",
                        f"DE 1/2 equals {add_type}.",
                        f"DE 1/10 equals {proc_code}.",
                        f"The combination derives procedure category {category}.",
                        "All goods items on the declaration derive the same category.",
                    ],
                    verification_steps=[
                        "Populate DE 1/1, 1/2 and 1/10 with the stated combination.",
                        "Confirm the mapping against the HMRC derivation matrix.",
                        "Submit a declaration with only combinations that derive one category.",
                        "Repeat with a mismatched combination to confirm DMSREJ behaviour.",
                    ],
                    extra={"derived_category": category},
                )
            )
            next_id += 1

    return rules, next_id


def parse_export_derivation_rules(source: Source, next_id: int) -> tuple[list[Rule], int]:
    workbook = load_workbook(source.local_path, read_only=True, data_only=True)
    ws = workbook["Procedure Category Derivation 1"]

    procedure_codes = [compact(ws.cell(4, c).value) for c in range(1, ws.max_column + 1)]
    rules: list[Rule] = []
    current_decl_type = ""

    for row_idx in range(5, ws.max_row + 1):
        decl_type = compact(ws.cell(row_idx, 2).value)
        add_type = compact(ws.cell(row_idx, 3).value)
        if decl_type:
            current_decl_type = decl_type
        if not current_decl_type or not add_type:
            continue
        for col_idx in range(4, ws.max_column + 1):
            proc_code = procedure_codes[col_idx - 1]
            category = compact(ws.cell(row_idx, col_idx).value)
            if not proc_code or not category or category == "-":
                continue
            desc = (
                f"Export procedure derivation: declaration type {current_decl_type}, additional declaration type "
                f"{add_type}, and requested procedure {proc_code} must derive category {category}."
            )[:255]
            rules.append(
                Rule(
                    sequential_id=next_id,
                    traceability_id=f"TDR-R-{next_id:05d}",
                    category="Cross-field integrity",
                    scope="Export declaration",
                    description=desc,
                    hmrc_reference=f"{source.title} | {source.version} | Procedure Category Derivation 1!{row_idx}",
                    source_id=source.source_id,
                    source_title=source.title,
                    source_version=source.version,
                    page_or_section=f"Procedure Category Derivation 1 row {row_idx}",
                    url=source.url,
                    last_updated=source.last_updated,
                    error_code="CDS11004 | CDS11005 | CDS12050 | CDS12052",
                    acceptance_criteria=[
                        f"DE 1/1 equals {current_decl_type}.",
                        f"DE 1/2 equals {add_type}.",
                        f"DE 1/10 equals {proc_code}.",
                        f"The combination derives procedure category {category}.",
                    ],
                    verification_steps=[
                        "Populate DE 1/1, 1/2 and 1/10 with the stated export combination.",
                        "Confirm the mapping against the HMRC export derivation matrix.",
                        "Submit an export declaration using a single derived category.",
                        "Introduce an invalid combination to confirm rejection handling.",
                    ],
                    extra={"derived_category": category},
                )
            )
            next_id += 1

    return rules, next_id


def build_manual_rules(sources: dict[str, Source], next_id: int) -> tuple[list[Rule], int]:
    rules: list[Rule] = []

    def add(
        category: str,
        scope: str,
        description: str,
        source_id: str,
        page_or_section: str,
        error_code: str,
        acceptance: list[str],
        steps: list[str],
        extra: dict[str, str] | None = None,
    ) -> None:
        nonlocal next_id
        source = sources[source_id]
        rules.append(
            Rule(
                sequential_id=next_id,
                traceability_id=f"TDR-R-{next_id:05d}",
                category=category,
                scope=scope,
                description=description[:255],
                hmrc_reference=f"{source.title} | {source.version} | {page_or_section}",
                source_id=source.source_id,
                source_title=source.title,
                source_version=source.version,
                page_or_section=page_or_section,
                url=source.url,
                last_updated=source.last_updated,
                error_code=error_code,
                acceptance_criteria=acceptance,
                verification_steps=steps,
                extra=extra or {},
            )
        )
        next_id += 1

    add(
        "Global validation",
        "All CDS XML messages",
        "Declaration, amendment, cancellation and goods-arrival payloads must pass HMRC XSD schema validation before CDS processing begins.",
        "dssd",
        "Section 2.2 Technical Validation Carried out on all Declarations",
        "BAD_REQUEST | xml_validation_error",
        [
            "Payload conforms to the HMRC-published XSD for the API and version in use.",
            "No datatype, element-order or invalid-element schema errors are returned.",
            "HTTP 400 is not returned by the API gateway.",
        ],
        [
            "Validate the payload against the relevant HMRC XSD.",
            "Submit the payload with the correct Accept header.",
            "Confirm that no BAD_REQUEST or xml_validation_error response is returned.",
        ],
    )
    add(
        "Global validation",
        "All CDS XML messages",
        "Payload content must not trigger the HMRC web application firewall regular-expression rules.",
        "dssd",
        "Section 2.2 Web Application Firewall (WAF) Validation",
        "HTTP 403 PAYLOAD_FORBIDDEN | HTTP 500 INTERNAL_SERVER_ERROR",
        [
            "Free-text fields do not contain blocked character combinations or keyword patterns.",
            "The API does not return PAYLOAD_FORBIDDEN or Internal server error at WAF stage.",
        ],
        [
            "Populate the message with production-like free text.",
            "Check the text against HMRC WAF examples and recommendations.",
            "Submit the payload and confirm that WAF does not reject it.",
        ],
    )
    add(
        "Global validation",
        "Declaration submission",
        "CDS does not accept batch declaration submission; declarations must be submitted individually.",
        "dssd",
        "Section 2 Scope for Declaration Management",
        "DMSREJ validationResultType | API rejection",
        [
            "Each API call contains one declaration message only.",
            "No batch wrapper or multi-declaration payload is sent.",
        ],
        [
            "Review the integration request builder.",
            "Confirm one declaration per API request.",
            "Submit and verify normal asynchronous processing.",
        ],
    )
    add(
        "Global validation",
        "Declaration submission",
        "Local Reference Number (LRN / submitterReference / DE 2/5) must be unique against accepted and registered declarations and cannot be reused within 48 hours.",
        "error_codes",
        "CDS12003",
        "CDS12003",
        [
            "DE 2/5 is unique for the trader among accepted and registered declarations.",
            "A rejected declaration's LRN may be reused; an accepted or registered one may not be reused within 48 hours.",
        ],
        [
            "Submit a declaration with a new LRN.",
            "Resubmit the same LRN within 48 hours for the same trader.",
            "Confirm HMRC returns CDS12003 for the duplicate submission.",
        ],
    )
    add(
        "Cross-field integrity",
        "Declaration submission",
        "DE 1/9 Total Number of Items must equal the number of goods items declared.",
        "error_codes",
        "CDS11003",
        "CDS11003",
        [
            "DE 1/9 matches the count of goods-item structures in the payload.",
            "No extra or missing goods-item nodes exist relative to DE 1/9.",
        ],
        [
            "Build a declaration with a known item count.",
            "Set DE 1/9 to the correct number and submit.",
            "Repeat with a mismatched value to confirm CDS11003.",
        ],
    )
    add(
        "Cross-field integrity",
        "Import and export declarations",
        "All goods items on a declaration must derive the same procedure category.",
        "error_codes",
        "CDS11005",
        "CDS11005",
        [
            "Every item's DE 1/10 and DE 1/2 combination derives the same declaration category.",
            "No mixed-category declaration is submitted.",
        ],
        [
            "Derive the category for every item using the HMRC derivation matrix.",
            "Confirm all items resolve to one category before submission.",
            "Submit a mixed-category declaration to confirm CDS11005 is returned.",
        ],
    )
    add(
        "Cross-field integrity",
        "Import and export declarations",
        "Net mass must be lower than gross mass.",
        "error_codes",
        "CDS11006",
        "CDS11006",
        ["Where both masses are declared, net mass is less than gross mass for the same scope."],
        [
            "Populate gross mass and net mass for the same item or header.",
            "Check that net mass is lower than gross mass.",
            "Submit a failing case to confirm CDS11006.",
        ],
    )
    add(
        "Additional-message validation",
        "Goods arrival",
        "A goods-arrival notification must follow a pre-lodged declaration in a permissible state.",
        "error_codes",
        "CDS12014",
        "CDS12014 | CDS12015",
        [
            "A pre-lodged declaration exists for the reference used by the goods-arrival message.",
            "The referenced declaration is in a state that permits arrival.",
        ],
        [
            "Submit a pre-lodged declaration.",
            "Submit a goods-arrival message for that declaration.",
            "Repeat with a non-pre-lodged or invalid-state reference to confirm CDS12014 or CDS12015.",
        ],
    )
    add(
        "Additional-message validation",
        "Goods arrival",
        "Goods-arrival notifications must be submitted within 30 days of the pre-lodged declaration.",
        "error_codes",
        "CDS12031 / CDS12046",
        "CDS12031 | CDS12046",
        ["Arrival is submitted within the 30-day window measured from the original declaration receipt date."],
        [
            "Record the receipt date of the pre-lodged declaration.",
            "Submit arrival within 30 days and confirm acceptance.",
            "Submit after 30 days to confirm CDS12031 or CDS12046.",
        ],
    )
    add(
        "Additional-message validation",
        "Amendment and cancellation",
        "A new amendment or cancellation request cannot be submitted while a previous request is still awaiting response.",
        "error_codes",
        "CDS12036",
        "CDS12036",
        ["Only one in-flight amendment or cancellation request exists per declaration lifecycle point."],
        [
            "Submit an amendment or cancellation request.",
            "Before a response arrives, submit another request.",
            "Confirm HMRC returns CDS12036.",
        ],
    )
    add(
        "Additional-message validation",
        "Amendment",
        "The submission channel of an amendment must match the original declaration channel.",
        "error_codes",
        "CDS12034",
        "CDS12034",
        ["The amendment uses the same submission channel as the original declaration."],
        [
            "Record the channel used for the original declaration.",
            "Submit the amendment through the same channel.",
            "Submit through a different channel to confirm CDS12034.",
        ],
    )
    add(
        "Format rule",
        "Supporting-document upload",
        "The secure file-upload service accepts at most 10 files, each up to 10 MB, and only jpeg, pdf, doc, docx, xls, xlsx, or png formats.",
        "file_upload_ui",
        "Section 3.1.5 File Upload",
        "HTTP 413 | BAD_REQUEST",
        [
            "No more than 10 files are uploaded in one submission set.",
            "Each file is 10 MB or smaller.",
            "Each file extension is one of jpeg, pdf, doc, docx, xls, xlsx, png.",
        ],
        [
            "Prepare a set of supporting-document files.",
            "Confirm file count, size, and extension before upload.",
            "Upload the files and verify a document reference code is returned for each accepted file.",
        ],
    )
    add(
        "Format rule",
        "Supporting-document upload",
        "Avoid %, ?, &, tabs, and other special characters in secure-file-upload filenames because the service may behave unexpectedly.",
        "file_upload_ui",
        "Section 3.1.5 File Upload",
        "BAD_REQUEST | HTTP 403 PAYLOAD_FORBIDDEN | Unexpected file-read failure",
        ["Filename characters are HTTP-safe and do not contain %, ?, &, tabs, or special characters flagged by HMRC."],
        [
            "Inspect each filename before upload.",
            "Rename files containing reserved or special characters.",
            "Upload and confirm successful document-reference issuance.",
        ],
    )
    add(
        "Format rule",
        "WAF-sensitive free-text fields",
        "Avoid semicolons, equals signs, slash characters in goods descriptions, and SQL-keyword patterns in free text because these combinations are known WAF triggers.",
        "waf_examples",
        "General Advice tab (GA1-GA4)",
        "HTTP 403 PAYLOAD_FORBIDDEN | HTTP 500 INTERNAL_SERVER_ERROR",
        [
            "Free-text fields do not contain ';', '=', SQL-keyword strings, or '/' in goodsDescription where HMRC has flagged them as trigger patterns.",
            "Known trigger examples are removed or rewritten before submission.",
        ],
        [
            "Review all free-text fields in the XML payload.",
            "Compare them against the WAF Trigger Examples workbook.",
            "Replace or rewrite flagged strings before submission.",
        ],
    )
    add(
        "Format rule",
        "Export inventory linking",
        "MUCR values must match HMRC's published CDS MUCR formats, contain no more than one hyphen, and MRNs must not be submitted as export movements.",
        "mucr_format",
        "Sections 2 and 4 Appendix",
        "MUCR reference is invalid | DMSREJ / ILE NAK",
        [
            "MUCR length is 35 characters or fewer.",
            "MUCR matches one HMRC-published CDS format.",
            "No MRN is used in place of a MUCR on export movement messaging.",
        ],
        [
            "Generate MUCRs from one of the HMRC-published format families.",
            "Check length and hyphen count before movement submission.",
            "Submit a failing format to confirm rejection behaviour in exports inventory linking.",
        ],
    )
    add(
        "Format rule",
        "Authorisations",
        "Authorisation references declared in DE 2/3 and DE 3/39 must match HMRC-published CDS authorisation formats for the declared authorisation type.",
        "validation_authorisation",
        "Authorisations Table",
        "CDS12005 | CDS12006 | CDS12007",
        [
            "The declared authorisation or licence reference matches the HMRC format for its type.",
            "The reference belongs to the declared party and is valid for the scenario.",
        ],
        [
            "Identify the authorisation type being declared.",
            "Validate the reference format against HMRC's authorisation guidance.",
            "Submit a declaration with an invalid or mismatched reference to confirm the relevant CDS1200x error.",
        ],
    )
    add(
        "Business-rule validation",
        "Tariff measures",
        "Where a tariff measure requires a document, CDS returns CDS40045 if the required document is missing; where a measure condition or range is not met, CDS returns CDS40066.",
        "dssd",
        "Section 6.3 Tariff Measure Validation Failures",
        "CDS40045 | CDS40066",
        [
            "All document-based measure conditions are declared on the affected commodity lines.",
            "All value, ratio, or other measure conditions are satisfied for the affected commodity lines.",
        ],
        [
            "Retrieve the applicable tariff measures for the commodity/origin combination.",
            "Populate all required documents and measure-condition data.",
            "Submit a negative test without the condition to confirm CDS40045 or CDS40066.",
        ],
    )
    add(
        "Integration rule",
        "Trader Dress Rehearsal environment",
        "The Accept header must target the TDR API version for each subscribed CDS API, including Customs Declarations API v1.0 in TDR.",
        "web_dev_setup",
        "API versioning",
        "HTTP 406 Not Acceptable",
        [
            "The request Accept header matches the TDR API version subscribed for that environment.",
            "The application is allow-listed and subscribed to that API version.",
        ],
        [
            "Inspect the outbound request headers.",
            "Confirm the application subscription for the TDR version.",
            "Submit a request with an incorrect Accept header to confirm HTTP 406 behaviour.",
        ],
    )
    add(
        "Operational rule",
        "Trader Dress Rehearsal proving",
        "TDR should be used for realistic, day-in-the-life declaration proving with real account data, not for performance or exploratory technical testing.",
        "runbook",
        "Slide 17 Guidance / Slide 2 Connecting to TDR",
        "Support escalation / misuse risk",
        [
            "TDR test scenarios use realistic declarant data and declaration journeys.",
            "The proving pack excludes performance, load, or exploratory protocol testing in TDR.",
        ],
        [
            "Review the planned TDR scenarios with the declarant.",
            "Confirm the data set is realistic and authorised for TDR use.",
            "Run technical experimentation and load tests only in the appropriate non-TDR environment.",
        ],
    )
    add(
        "Operational rule",
        "Service availability",
        "Submission planning must account for HMRC's published service-availability notices, planned maintenance windows, and queued-processing warnings.",
        "web_service_availability",
        "Current and historical update entries",
        "Service queueing / delayed processing",
        [
            "Submission windows are checked against the HMRC service-availability page before go-live or dress-rehearsal runs.",
            "Exports needing Permission to Progress are not first submitted during a published queueing window.",
        ],
        [
            "Check the latest HMRC service-availability page before executing the test window.",
            "Record any planned maintenance or outage affecting exports or imports.",
            "Adjust the run schedule to avoid queueing or blackout periods.",
        ],
    )
    return rules, next_id


def build_best_practice_findings() -> dict[str, list[dict[str, str]]]:
    return {
        "data_preparation": [
            {
                "title": "Path-to-production workflow",
                "source": "The CDS Path to Production (page 5) and CDS RunBook slide 10",
                "detail": "Use Trade Test first, complete pre-TDR steps, register a TDR app, request TDR API access from SDST, subscribe, then practice declaration scenarios in TDR with real account data on an ongoing basis.",
            },
            {
                "title": "TDR proving approach",
                "source": "CDS RunBook slides 2, 17 and GOV.UK TDR guidance",
                "detail": "Use realistic declaration journeys and real account data; do not use TDR for performance or exploratory protocol testing.",
            },
            {
                "title": "Due diligence before submission",
                "source": "CDS RunBook slide 17",
                "detail": "Ensure all required data elements are populated correctly, check supporting-document needs, coordinate with supply-chain parties and CSPs, and review CDS error codes before escalating to HMRC.",
            },
        ],
        "xml_examples": [
            {
                "title": "Declaration submission message skeleton",
                "source": "CDS 03 DSSD v2.30 and WCO References workbook",
                "detail": "HMRC documents center declaration payloads on the WCO Declaration class with DE-driven paths such as Declaration/TypeCode, GovernmentAgencyGoodsItem/SequenceNumeric, and goods-item substructures. The reviewed public pack provides path definitions and fragments rather than a single standalone full canonical XML file.",
            },
            {
                "title": "Additional-message skeleton",
                "source": "CDS Additional Message Technical Completion Matrix v1.4",
                "detail": "Cancellation, amendment, and goods-arrival messages all require Declaration/FunctionalReferenceID, Declaration/ID and Declaration/TypeCode, with additional information blocks used for amendment and cancellation reasons.",
            },
            {
                "title": "Notification metadata and pointer structure",
                "source": "CDS 03 DSSD v2.30 section 6.2/6.3",
                "detail": "HMRC notifications rely on Response, MetaData, Error, Pointer and AdditionalInformation objects. Pointer resolution uses DocumentSectionCode, SequenceNumeric and TagID to identify the failing element.",
            },
            {
                "title": "IE-message coverage",
                "source": "CDS External TT TDR Scope Guidance 2025-12-04 issue",
                "detail": "The reviewed scope bulletins reference IE501, IE510, IE518 and IE591 scenarios for XI/AES functionality, but those slides are mostly acceptance-criteria statements rather than complete public XML specimens.",
            },
        ],
        "error_blacklist": [
            {
                "title": "WAF-sensitive strings",
                "source": "WAF Trigger Examples v0.4",
                "detail": "Avoid semicolons, equals signs, SQL keywords and slash characters in goods descriptions; HMRC examples include 'VALVE - BODY ONLY ( = 1/2\")', '; SELECT WINES OF THE DORDOGNE' and bare slash usage in goodsDescription.",
            },
            {
                "title": "High-frequency validation families",
                "source": "CDS Error Codes 11-03-2026 issue",
                "detail": "The most reusable prevention buckets are missing mandatory data (CDS10001), invalid format (CDS10010), invalid domain value (CDS10020), invalid relation or conditional dependency (CDS11004, CDS12070-72), and tariff-measure conditions (CDS40045, CDS40066).",
            },
            {
                "title": "Known differences and workarounds",
                "source": "Declarant_CWD_Pre_5.0.0.ods",
                "detail": "The known-differences workbook remains necessary for edge cases where current CDS behaviour differs from tariff intent; declarants should check the document when a listed error code appears or when a workaround is mandated.",
            },
        ],
        "submission_timing": [
            {
                "title": "Maintenance windows",
                "source": "Customs Declaration Service: service availability and issues",
                "detail": "HMRC publishes planned maintenance windows and queueing notices on the service-availability page instead of a fixed daily TDR cut-off timetable. Example: exports submitted during the 18 February 2026 maintenance window were queued until completion.",
            },
            {
                "title": "Rate limiting",
                "source": "Developer set up guide",
                "detail": "Standard CDS rate limit is 3 requests per second per application, with CDS users pre-approved for 8 requests per second if obtained through SDST; applications should back off and retry after HTTP 429 responses.",
            },
            {
                "title": "Token window",
                "source": "CDS RunBook slide 12",
                "detail": "A declarant access token expires after 4 hours and needs refresh handling in TDR and production.",
            },
        ],
        "qa_checkpoints": [
            {
                "title": "Three-stage validation",
                "source": "CDS 03 DSSD v2.30",
                "detail": "Validate every message against the XSD, then against known WAF trigger patterns, then against business rules before treating a scenario as TDR-ready.",
            },
            {
                "title": "Header and version check",
                "source": "Developer set up guide and CDS RunBook slide 10",
                "detail": "Confirm the Accept header matches the TDR version for the subscribed API before each test run.",
            },
            {
                "title": "File-upload hygiene",
                "source": "CDS Secure File Upload UI User Guide v0.44",
                "detail": "Validate file count, size, extension and filename safety before upload; successful uploads should return document-reference codes that can be retained as evidence.",
            },
        ],
    }


def build_source_notes() -> list[dict[str, str]]:
    return [
        {
            "title": "Official publication inventory scope",
            "detail": "The reviewed source set combines public GOV.UK guidance pages with the HMRC Developer Hub technical-documentation bundle that HMRC states is the latest issued pack for external software developers.",
        },
        {
            "title": "Historical-update coverage",
            "detail": "The change-log document dated 20 March 2026 was used as the anchor for identifying documents updated since the TDR pilot period began. Individual internal files do not always expose a stand-alone public URL; the audit therefore cites the bundle URL plus the internal file path or worksheet/slide identifier.",
        },
        {
            "title": "Evidence gaps",
            "detail": "HMRC does not publish a public consolidated validation-rule count for TDR, nor does it publish traders' zero-error confirmation emails or screenshots. Those quality-gate evidence items remain private-programme artifacts outside the public corpus.",
        },
        {
            "title": "Public-vs-gated documentation",
            "detail": "The DSSD references API specifications and XSDs on the Developer Hub. Some API endpoints and example-message pages may require allow-listing or sign-in, but the reviewed public bundle still provides the core completion matrices, path definitions and validation semantics needed for this audit.",
        },
    ]


def build_use_case_findings() -> list[dict[str, str]]:
    return [
        {
            "source": "The Bonded Warehousekeepers Association - HMRC Webinar: CDS - Trader Dress Rehearsal",
            "url": "https://www.thebwa.com/hmrc-webinar-cds-trader-dress-rehearsal/",
            "date": "2021-09-15",
            "relevance_score": "5/5",
            "summary": "Trade-association post describing an HMRC-led overview session focused on TDR benefits, migration planning and support routes; useful as externally hosted evidence of HMRC-endorsed TDR onboarding messages.",
        },
        {
            "source": "The Bonded Warehousekeepers Association - CDS Trader Dress Rehearsal Declarant Roundtable (PDF)",
            "url": "https://www.thebwa.com/wp-content/uploads/CDS-Trader-Dress-Rehearsal-Declarant-Roundtable.pdf",
            "date": "2021-09-01",
            "relevance_score": "5/5",
            "summary": "Externally hosted but HMRC-branded 'OFFICIAL' slide deck covering declarant-focused TDR positioning, migration benefits, and support expectations.",
        },
        {
            "source": "Charles Kendall / HMRC Readiness Presentation (PDF mirror)",
            "url": "https://www.charleskendall.com/uploads/files/CDS-HMRC-Readiness-Presentation-18-May-2022.pdf",
            "date": "2022-05-18",
            "relevance_score": "4/5",
            "summary": "Mirror of an HMRC town-hall deck that explicitly includes the TDR stage and repeats key programme warnings such as 'not for performance testing' and 'doesn't create legal declarations'.",
        },
        {
            "source": "ICAEW article linking to HMRC export migration webinar",
            "url": "https://www.icaew.com/insights/tax-news/2024/mar-2024/uk-exporters-must-switch-to-cds-by-june",
            "date": "2024-03-04",
            "relevance_score": "3/5",
            "summary": "Professional-body summary that points traders to HMRC's free TDR service and a separate HMRC webinar on preparing to move exports to CDS; useful corroboration but secondary to official HMRC sources.",
        },
        {
            "source": "Air Cargo Week report on BIFA / ASM migration webinar",
            "url": "https://ppd.aircargoweek.com/bifa-and-asm-webinar-prepares-businesses-for-final-migration-to-new-customs-declaration-service-for-exports/",
            "date": "2024-01-26",
            "relevance_score": "3/5",
            "summary": "Trade-press write-up reporting that BIFA and ASM highlighted higher migration success among TDR users; relevant as industry evidence but not a normative HMRC rule source.",
        },
    ]


def build_change_log_summary(source: Source) -> list[str]:
    lines = odf_text(Path(source.local_path))
    keep = []
    for line in lines:
        low = line.lower()
        if any(
            token in low
            for token in [
                "cds 03 declaration tcm",
                "cds 03 cds codelists",
                "scope guidance",
                "known error",
                "runbook",
                "pre 5.0.0",
                "trade_test_kel",
            ]
        ):
            keep.append(line)
    return keep[:40]


def build_waf_summary(source: Source) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    workbook = load_workbook(source.local_path, read_only=True, data_only=True)
    general_ws = workbook["General Advice"]
    example_ws = workbook["Examples"]
    advice = []
    examples = []
    for row in general_ws.iter_rows(min_row=3, values_only=True):
        rid, recommendation, reasoning, added = row[1:5]
        if rid:
            advice.append(
                {
                    "id": compact(rid),
                    "recommendation": compact(recommendation),
                    "reasoning": compact(reasoning),
                    "added": compact(added),
                }
            )
    for row in example_ws.iter_rows(min_row=3, values_only=True):
        rid, example, first_seen, last_tested, recommendation, added = row[1:7]
        if rid:
            examples.append(
                {
                    "id": compact(rid),
                    "example": compact(example),
                    "first_seen": first_seen.strftime("%Y-%m-%d") if isinstance(first_seen, datetime) else compact(first_seen),
                    "last_tested": last_tested.strftime("%Y-%m-%d") if isinstance(last_tested, datetime) else compact(last_tested),
                    "recommendation": compact(recommendation),
                    "added": compact(added),
                }
            )
    return advice, examples


def build_known_difference_summary(source: Source) -> dict[str, list[str]]:
    xl = pd.ExcelFile(source.local_path, engine="odf")
    summary: dict[str, list[str]] = {}
    for sheet in ["KD_Change_Log", "KEL_Change_Log"]:
        df = xl.parse(sheet, header=None)
        lines = []
        for idx in range(min(len(df), 10)):
            left = compact(df.iloc[idx, 0])
            right = compact(df.iloc[idx, 1] if df.shape[1] > 1 else "")
            if left or right:
                lines.append(f"{left}: {right}".strip(": "))
        summary[sheet] = lines
    return summary


def build_rule_dataset() -> dict[str, object]:
    sources = build_sources()
    error_code_map = build_error_code_map(Path(sources["error_codes"].local_path))

    next_id = 1
    import_fields, import_rules, next_id = parse_import_field_rules(sources["imports_tcm"], next_id)
    export_fields, export_rules, next_id = parse_export_field_rules(sources["exports_tcm"], next_id)
    additional_fields, additional_rules, next_id = parse_additional_message_rules(
        sources["additional_message_tcm"], next_id
    )
    import_derivation_rules, next_id = parse_import_derivation_rules(sources["imports_tcm"], next_id)
    export_derivation_rules, next_id = parse_export_derivation_rules(sources["exports_tcm"], next_id)
    manual_rules, next_id = build_manual_rules(sources, next_id)

    rules = (
        manual_rules
        + import_derivation_rules
        + export_derivation_rules
        + additional_rules
        + import_rules
        + export_rules
    )
    rules = sorted(rules, key=lambda r: r.sequential_id)
    for idx, rule in enumerate(rules, start=1):
        rule.sequential_id = idx
        rule.traceability_id = f"TDR-R-{idx:05d}"

    code_list_counts = Counter(
        compact(entry.get("code_list"))
        for entry in [*import_fields, *export_fields, *additional_fields]
        if compact(entry.get("code_list"))
    )
    summary = {
        "rules_total": len(rules),
        "field_rules_total": len(import_rules) + len(export_rules) + len(additional_rules),
        "cross_field_rules_total": len(import_derivation_rules) + len(export_derivation_rules),
        "manual_rules_total": len(manual_rules),
        "import_field_count": len(import_fields),
        "export_field_count": len(export_fields),
        "additional_message_field_count": len(additional_fields),
        "top_code_lists": code_list_counts.most_common(25),
    }

    return {
        "sources": sources,
        "rules": rules,
        "error_code_map": error_code_map,
        "field_catalog": {
            "imports": import_fields,
            "exports": export_fields,
            "additional_messages": additional_fields,
        },
        "best_practice": build_best_practice_findings(),
        "use_cases": build_use_case_findings(),
        "source_notes": build_source_notes(),
        "summary": summary,
        "waf_advice": build_waf_summary(sources["waf_examples"])[0],
        "waf_examples": build_waf_summary(sources["waf_examples"])[1],
        "known_diff_summary": build_known_difference_summary(sources["known_differences"]),
        "change_log_summary": build_change_log_summary(sources["tech_change_log"]),
    }


def export_rule_csv(rules: list[Rule], output_path: Path) -> None:
    rows = []
    for rule in rules:
        rows.append(
            {
                "Sequential ID": rule.sequential_id,
                "Traceability ID": rule.traceability_id,
                "Category": rule.category,
                "Scope": rule.scope,
                "Rule description": rule.description,
                "HMRC reference": rule.hmrc_reference,
                "Publication title": rule.source_title,
                "Version": rule.source_version,
                "Page/paragraph identifier": rule.page_or_section,
                "URL": rule.url,
                "Last updated": rule.last_updated,
                "Error code": rule.error_code,
                "Acceptance criteria": "\n".join(f"- {item}" for item in rule.acceptance_criteria),
                "Verification procedure": "\n".join(
                    f"{idx}. {item}" for idx, item in enumerate(rule.verification_steps, start=1)
                ),
            }
        )
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def export_checklist_workbook(rules: list[Rule], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "TDR Checklist"

    headers = [
        "Sequential ID (integer, unique, gap-free)",
        "Rule description (≤ 255 characters, plain English)",
        "HMRC reference (concatenated string: “title | version | page”)",
        "Error code",
        "Acceptance criteria (bullet-point list)",
        "Step-by-step verification procedure (numbered steps)",
        "Pass/fail status (dropdown: Pass / Fail / Not-tested)",
        "Evidence attachment link (hyperlink to screenshot, XML snippet, or test-report PDF)",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for rule in rules:
        ws.append(
            [
                rule.sequential_id,
                rule.description[:255],
                rule.hmrc_reference,
                rule.error_code,
                "\n".join(f"• {item}" for item in rule.acceptance_criteria),
                "\n".join(f"{idx}. {item}" for idx, item in enumerate(rule.verification_steps, start=1)),
                rule.pass_fail_status,
                rule.evidence_link,
            ]
        )

    dv = DataValidation(type="list", formula1='"Pass,Fail,Not-tested"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"G2:G{ws.max_row}")

    widths = {"A": 12, "B": 48, "C": 46, "D": 28, "E": 52, "F": 58, "G": 18, "H": 42}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    wb.save(output_path)


def iso_dt(d: date, end: bool = False) -> str:
    base = datetime.combine(d, time(17, 0) if end else time(9, 0))
    return base.strftime("%Y-%m-%dT%H:%M:%S")


def export_project_xml(output_path: Path) -> None:
    t0 = date(2026, 3, 26)
    tasks = [
        {"uid": 1, "id": 1, "name": "HMRC Trader Dress Rehearsal Audit Programme", "outline": 1, "summary": 1, "start": iso_dt(t0), "finish": iso_dt(t0 + timedelta(days=32), end=True), "notes": "Programme summary task for the public-documentation audit."},
        {"uid": 2, "id": 2, "name": "WP-100 Documentation harvesting", "outline": 2, "summary": 0, "start": iso_dt(t0), "finish": iso_dt(t0 + timedelta(days=10), end=True), "resource": "Business Analyst", "notes": "Collect official HMRC/GOV.UK publications, schemas, matrices, release notes, bulletins and change logs relevant to TDR."},
        {"uid": 3, "id": 3, "name": "WP-200 Rule extraction & mapping", "outline": 2, "summary": 0, "start": iso_dt(t0 + timedelta(days=11)), "finish": iso_dt(t0 + timedelta(days=18), end=True), "resource": "Solution Architect", "pred": 2, "notes": "Normalize rules, trace sources, and map validation logic into a structured register."},
        {"uid": 4, "id": 4, "name": "Gate-1 Evidence Review", "outline": 2, "summary": 0, "start": iso_dt(t0 + timedelta(days=18)), "finish": iso_dt(t0 + timedelta(days=18), end=True), "milestone": 1, "resource": "Solution Architect", "pred": 3, "notes": "Pass criteria: 100% of compulsory rules identified and documented. Evidence: checklist row count equals the HMRC validation-rule count or the closest public-source proxy count where HMRC has not published a single consolidated count. Public-source caveat must be logged if a complete HMRC count is unavailable."},
        {"uid": 5, "id": 5, "name": "WP-300 Checklist build", "outline": 2, "summary": 0, "start": iso_dt(t0 + timedelta(days=19)), "finish": iso_dt(t0 + timedelta(days=22), end=True), "resource": "QA Lead", "pred": 4, "notes": "Generate the traceable test checklist workbook and align the rule CSV and PDF appendix."},
        {"uid": 6, "id": 6, "name": "WP-400 Internal peer review", "outline": 2, "summary": 0, "start": iso_dt(t0 + timedelta(days=23)), "finish": iso_dt(t0 + timedelta(days=26), end=True), "resource": "HMRC Subject-Matter Expert", "pred": 5, "notes": "Review traceability, testability and public-source alignment."},
        {"uid": 7, "id": 7, "name": "Gate-2 QA Sign-off", "outline": 2, "summary": 0, "start": iso_dt(t0 + timedelta(days=26)), "finish": iso_dt(t0 + timedelta(days=26), end=True), "milestone": 1, "resource": "QA Lead", "pred": 6, "notes": "Pass criteria: 100% of acceptance criteria are testable. Evidence: QA Lead review sign-off on the checklist wording and procedures."},
        {"uid": 8, "id": 8, "name": "WP-500 Dry-run validation", "outline": 2, "summary": 0, "start": iso_dt(t0 + timedelta(days=27)), "finish": iso_dt(t0 + timedelta(days=30), end=True), "resource": "Test Manager", "pred": 7, "notes": "Execute the checklist against representative historical or rehearsal submissions and collect evidence."},
        {"uid": 9, "id": 9, "name": "Gate-3 Historical zero-error comparison", "outline": 2, "summary": 0, "start": iso_dt(t0 + timedelta(days=30)), "finish": iso_dt(t0 + timedelta(days=30), end=True), "milestone": 1, "resource": "Test Manager", "pred": 8, "notes": "Pass criteria: checklist validated against a minimum of three historical TDR submissions that achieved zero HMRC errors. Evidence: hyperlinks to zero-error confirmation emails or HMRC portal screenshots. This evidence is private-operational and not available in the public HMRC corpus; acquisition is a programme dependency."},
        {"uid": 10, "id": 10, "name": "WP-600 Final sign-off", "outline": 2, "summary": 0, "start": iso_dt(t0 + timedelta(days=31)), "finish": iso_dt(t0 + timedelta(days=32), end=True), "resource": "Programme Director", "pred": 9, "notes": "Approve final issue pack and delivery to stakeholders."},
    ]
    resources = [(1, "Business Analyst"), (2, "Solution Architect"), (3, "QA Lead"), (4, "HMRC Subject-Matter Expert"), (5, "Test Manager"), (6, "Programme Director")]
    resource_uid = {name: uid for uid, name in resources}

    def duration_hours(start_s: str, finish_s: str) -> str:
        start_dt = datetime.strptime(start_s, "%Y-%m-%dT%H:%M:%S")
        finish_dt = datetime.strptime(finish_s, "%Y-%m-%dT%H:%M:%S")
        hours = int(max((finish_dt - start_dt).total_seconds() // 3600, 0))
        return f"PT{hours}H0M0S"

    task_xml = []
    for task in tasks:
        pieces = [
            "<Task>", f"<UID>{task['uid']}</UID>", f"<ID>{task['id']}</ID>", f"<Name>{escape(task['name'])}</Name>", "<Type>1</Type>", "<IsNull>0</IsNull>", f"<CreateDate>{iso_dt(t0)}</CreateDate>", f"<WBS>{task['id']}</WBS>", f"<OutlineNumber>{task['id']}</OutlineNumber>", f"<OutlineLevel>{task['outline']}</OutlineLevel>", "<Priority>500</Priority>", f"<Start>{task['start']}</Start>", f"<Finish>{task['finish']}</Finish>", f"<Duration>{duration_hours(task['start'], task['finish'])}</Duration>", f"<Summary>{task.get('summary', 0)}</Summary>", f"<Milestone>{task.get('milestone', 0)}</Milestone>", f"<Notes>{escape(task['notes'])}</Notes>",
        ]
        if task.get("pred"):
            pieces.extend(["<PredecessorLink>", f"<PredecessorUID>{task['pred']}</PredecessorUID>", "<Type>1</Type>", "</PredecessorLink>"])
        pieces.append("</Task>")
        task_xml.append("".join(pieces))

    resource_xml = [
        "".join(["<Resource>", f"<UID>{uid}</UID>", f"<ID>{uid}</ID>", f"<Name>{escape(name)}</Name>", "<Type>1</Type>", "</Resource>"])
        for uid, name in resources
    ]
    assignment_xml = []
    assign_uid = 1
    for task in tasks:
        resource_name = task.get("resource")
        if not resource_name:
            continue
        assignment_xml.append(
            "".join(
                [
                    "<Assignment>",
                    f"<UID>{assign_uid}</UID>",
                    f"<TaskUID>{task['uid']}</TaskUID>",
                    f"<ResourceUID>{resource_uid[resource_name]}</ResourceUID>",
                    "<Units>1</Units>",
                    "</Assignment>",
                ]
            )
        )
        assign_uid += 1

    xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<Project xmlns="http://schemas.microsoft.com/project">
  <SaveVersion>14</SaveVersion>
  <Name>HMRC Trader Dress Rehearsal Audit Programme</Name>
  <Title>HMRC Trader Dress Rehearsal Audit Programme</Title>
  <CreationDate>{iso_dt(t0)}</CreationDate>
  <ScheduleFromStart>1</ScheduleFromStart>
  <StartDate>{iso_dt(t0)}</StartDate>
  <DefaultStartTime>09:00:00</DefaultStartTime>
  <DefaultFinishTime>17:00:00</DefaultFinishTime>
  <MinutesPerDay>480</MinutesPerDay>
  <MinutesPerWeek>2400</MinutesPerWeek>
  <DaysPerMonth>20</DaysPerMonth>
  <Tasks>{''.join(task_xml)}</Tasks>
  <Resources>{''.join(resource_xml)}</Resources>
  <Assignments>{''.join(assignment_xml)}</Assignments>
</Project>
"""
    output_path.write_text(xml, encoding="utf-8")


def export_source_inventory(sources: dict[str, Source], output_path: Path) -> None:
    output_path.write_text(json.dumps([asdict(s) for s in sources.values()], indent=2), encoding="utf-8")


def export_raw_rules(rules: list[Rule], output_path: Path) -> None:
    output_path.write_text(json.dumps([asdict(r) for r in rules], indent=2), encoding="utf-8")


def draw_page_number(canvas, doc) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 9)
    canvas.drawRightString(195 * mm, 10 * mm, f"Page {doc.page}")
    canvas.restoreState()


def paragraph(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape(text).replace("\n", "<br/>"), style)


def export_master_pdf(dataset: dict[str, object], output_path: Path) -> None:
    sources: dict[str, Source] = dataset["sources"]  # type: ignore[assignment]
    rules: list[Rule] = dataset["rules"]  # type: ignore[assignment]
    field_catalog: dict[str, list[dict[str, str]]] = dataset["field_catalog"]  # type: ignore[assignment]
    best_practice: dict[str, list[dict[str, str]]] = dataset["best_practice"]  # type: ignore[assignment]
    use_cases: list[dict[str, str]] = dataset["use_cases"]  # type: ignore[assignment]
    source_notes: list[dict[str, str]] = dataset["source_notes"]  # type: ignore[assignment]
    summary: dict[str, object] = dataset["summary"]  # type: ignore[assignment]
    waf_advice: list[dict[str, str]] = dataset["waf_advice"]  # type: ignore[assignment]
    waf_examples: list[dict[str, str]] = dataset["waf_examples"]  # type: ignore[assignment]
    known_diff_summary: dict[str, list[str]] = dataset["known_diff_summary"]  # type: ignore[assignment]
    change_log_summary: list[str] = dataset["change_log_summary"]  # type: ignore[assignment]

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleSmall", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=20, leading=24, alignment=TA_CENTER, spaceAfter=8))
    styles.add(ParagraphStyle(name="HeadingMain", parent=styles["Heading1"], fontName="Helvetica-Bold", fontSize=14, leading=18, spaceBefore=10, spaceAfter=6))
    styles.add(ParagraphStyle(name="HeadingSub", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=11, leading=14, spaceBefore=8, spaceAfter=4))
    styles.add(ParagraphStyle(name="BodyTight", parent=styles["BodyText"], fontName="Helvetica", fontSize=9, leading=11, spaceAfter=3, alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="RuleStyle", parent=styles["BodyText"], fontName="Helvetica", fontSize=8, leading=10, spaceAfter=2))

    story = []
    story.append(paragraph("HMRC Trader Dress Rehearsal (TDR) Master Requirements Document", styles["TitleSmall"]))
    story.append(paragraph("Public-source audit pack generated from HMRC and GOV.UK publications reviewed on 2026-03-26.", styles["BodyTight"]))
    story.append(paragraph("This document consolidates the official HMRC publications governing the Trader Dress Rehearsal programme that were available in the reviewed public corpus. It includes rule extraction, source traceability, best-practice consolidation, a change-log summary, and an evidence-gap statement for items HMRC does not publish publicly.", styles["BodyTight"]))
    story.append(Spacer(1, 6))

    story.append(paragraph("Audit Summary", styles["HeadingMain"]))
    summary_rows = [["Metric", "Value"], ["Total checklist rows / rules", str(summary["rules_total"])], ["Expanded field-level rules", str(summary["field_rules_total"])], ["Cross-field derivation rules", str(summary["cross_field_rules_total"])], ["Manual/global rules", str(summary["manual_rules_total"])], ["Unique import field definitions", str(summary["import_field_count"])], ["Unique export field definitions", str(summary["export_field_count"])], ["Unique additional-message field definitions", str(summary["additional_message_field_count"])]]
    table = LongTable(summary_rows, colWidths=[80 * mm, 80 * mm], repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), 0.3, colors.grey), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("FONTSIZE", (0, 0), (-1, -1), 8)]))
    story.append(table)
    story.append(Spacer(1, 6))

    story.append(paragraph("Source-Scope Notes", styles["HeadingMain"]))
    for note in source_notes:
        story.append(paragraph(f"{note['title']}: {note['detail']}", styles["BodyTight"]))

    story.append(PageBreak())
    story.append(paragraph("1. Mandatory Validation Rules", styles["HeadingMain"]))
    story.append(paragraph("HMRC validation for TDR submissions is layered. The DSSD describes three gates: XSD schema validation at the API gateway, WAF validation over the whole XML document, and CDS business-rule validation that returns asynchronous DMSREJ notifications with validationResultType codes.", styles["BodyTight"]))
    for rule in [r for r in rules if r.category in {"Global validation", "Business-rule validation", "Cross-field integrity", "Additional-message validation"}][:40]:
        story.append(paragraph(f"{rule.traceability_id} [{rule.category}] {rule.description} Reference: {rule.hmrc_reference}. Error pattern: {rule.error_code}.", styles["RuleStyle"]))

    story.append(paragraph("Recent Change-Log Highlights", styles["HeadingSub"]))
    for line in change_log_summary:
        story.append(paragraph(line, styles["RuleStyle"]))
    story.append(paragraph("WAF Preventive Rules", styles["HeadingSub"]))
    for item in waf_advice:
        story.append(paragraph(f"{item['id']}: {item['recommendation']} Reason: {item['reasoning']}. Added: {item['added']}.", styles["RuleStyle"]))
    story.append(paragraph("Known WAF Examples", styles["HeadingSub"]))
    for item in waf_examples[:8]:
        story.append(paragraph(f"{item['id']}: {item['example']} Recommendation: {item['recommendation']}. First seen: {item['first_seen']}.", styles["RuleStyle"]))

    story.append(PageBreak())
    story.append(paragraph("2. Data-Format Rules", styles["HeadingMain"]))
    story.append(paragraph("The TCM workbooks, codelists workbook, secure-file-upload guide, WAF bulletin, authorisation guidance and MUCR format note collectively govern data-format constraints. The public HMRC corpus does not publish a single flat table of every format rule, so this section consolidates the formats and references used by the reviewed documents.", styles["BodyTight"]))
    code_rows = [["Codelist", "Referenced fields in extracted catalog"]]
    for code_name, count in summary["top_code_lists"][:20]:  # type: ignore[index]
        code_rows.append([str(code_name), str(count)])
    code_table = LongTable(code_rows, colWidths=[120 * mm, 40 * mm], repeatRows=1)
    code_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.grey), ("FONTSIZE", (0, 0), (-1, -1), 8)]))
    story.append(code_table)
    story.append(Spacer(1, 6))
    for rule in [r for r in rules if r.category in {"Format rule", "Integration rule", "Operational rule"}]:
        story.append(paragraph(f"{rule.traceability_id}: {rule.description} Reference: {rule.hmrc_reference}. Error pattern: {rule.error_code}.", styles["RuleStyle"]))

    story.append(PageBreak())
    story.append(paragraph("3. Field-Level Requirements", styles["HeadingMain"]))
    story.append(paragraph("The following tables consolidate field metadata from the import, export and additional-message completion matrices. Each field retains its XML path, scope level, cardinality, data format, linked codelist (where identifiable from HMRC text), and conditional obligation matrix.", styles["BodyTight"]))

    def add_field_catalog(title: str, fields: list[dict[str, str]]) -> None:
        story.append(paragraph(title, styles["HeadingSub"]))
        display_fields = fields[: min(len(fields), 45)]
        rows = [["DE/Path", "Field", "XML path", "Lvl", "Card.", "Fmt", "Code list", "Obligations"]]
        for entry in display_fields:
            obligations = [f"{k.replace('obligation_', '')}={v}" for k, v in entry.items() if k.startswith("obligation_")]
            rows.append([entry.get("de", "") or entry.get("xml_path", ""), entry.get("field_name", ""), entry.get("xml_path", ""), entry.get("level", ""), entry.get("cardinality", ""), entry.get("format", ""), entry.get("code_list", ""), "; ".join(obligations[:8])])
        tbl = LongTable(rows, colWidths=[20 * mm, 30 * mm, 55 * mm, 10 * mm, 12 * mm, 16 * mm, 26 * mm, 40 * mm], repeatRows=1)
        tbl.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.grey), ("FONTSIZE", (0, 0), (-1, -1), 6.5), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
        story.append(tbl)
        if len(fields) > len(display_fields):
            story.append(paragraph(f"Only the first {len(display_fields)} rows are printed in the PDF section for readability. The full expanded rule set is embedded in the CSV/XLSX deliverables and the raw JSON export.", styles["RuleStyle"]))

    add_field_catalog("Import Field Catalog (sampled in PDF)", field_catalog["imports"])
    add_field_catalog("Export Field Catalog (sampled in PDF)", field_catalog["exports"])
    add_field_catalog("Additional-Message Field Catalog (sampled in PDF)", field_catalog["additional_messages"])

    story.append(PageBreak())
    story.append(paragraph("4. Best-Practice Compendium", styles["HeadingMain"]))
    for section_name, items in best_practice.items():
        story.append(paragraph(section_name.replace("_", " ").title(), styles["HeadingSub"]))
        for item in items:
            story.append(paragraph(f"{item['title']}: {item['detail']} Source: {item['source']}.", styles["BodyTight"]))
    story.append(paragraph("Known Difference Log Snapshot", styles["HeadingSub"]))
    for sheet, lines in known_diff_summary.items():
        story.append(paragraph(sheet, styles["RuleStyle"]))
        for line in lines:
            story.append(paragraph(line, styles["RuleStyle"]))

    story.append(PageBreak())
    story.append(paragraph("5. Source Inventory", styles["HeadingMain"]))
    source_rows = [["ID", "Title", "Version", "Last updated", "Type", "Reference method"]]
    for source in sources.values():
        source_rows.append([source.source_id, source.title, source.version, source.last_updated, source.publication_type, source.identifier_hint])
    source_table = LongTable(source_rows, colWidths=[22 * mm, 58 * mm, 24 * mm, 20 * mm, 30 * mm, 36 * mm], repeatRows=1)
    source_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.25, colors.grey), ("FONTSIZE", (0, 0), (-1, -1), 7)]))
    story.append(source_table)

    story.append(PageBreak())
    story.append(paragraph("6. Use-Case Supplementary Search", styles["HeadingMain"]))
    story.append(paragraph("A targeted search was run for externally hosted but HMRC-endorsed or HMRC-originated TDR implementation material, including trade-association webinar posts, mirrored HMRC slide decks and professional-body summaries.", styles["BodyTight"]))
    for item in use_cases:
        story.append(paragraph(f"Source: {item['source']}. Date: {item['date']}. Relevance: {item['relevance_score']}. Summary: {item['summary']} URL: {item['url']}", styles["BodyTight"]))

    story.append(PageBreak())
    story.append(paragraph("Appendix A. Rule Metadata Index", styles["HeadingMain"]))
    story.append(paragraph("The appendix prints a compact per-rule index. The CSV and checklist workbook contain the same rule set in machine-readable form.", styles["BodyTight"]))
    for rule in rules:
        story.append(paragraph(f"{rule.traceability_id} | {rule.category} | {rule.scope} | {rule.description}\nReference: {rule.hmrc_reference}\nError code/pattern: {rule.error_code}\nAcceptance: {'; '.join(rule.acceptance_criteria)}", styles["RuleStyle"]))

    doc = SimpleDocTemplate(str(output_path), pagesize=A4, leftMargin=12 * mm, rightMargin=12 * mm, topMargin=12 * mm, bottomMargin=14 * mm, title="HMRC Trader Dress Rehearsal Master Requirements Document", author="OpenAI Codex")
    doc.build(story, onFirstPage=draw_page_number, onLaterPages=draw_page_number)


def main() -> None:
    dataset = build_rule_dataset()
    sources: dict[str, Source] = dataset["sources"]  # type: ignore[assignment]
    rules: list[Rule] = dataset["rules"]  # type: ignore[assignment]

    export_rule_csv(rules, OUTPUT_DIR / "hmrc_tdr_rule_metadata.csv")
    export_checklist_workbook(rules, OUTPUT_DIR / "hmrc_tdr_traceable_test_checklist.xlsx")
    export_project_xml(OUTPUT_DIR / "hmrc_tdr_project_plan.xml")
    export_source_inventory(sources, OUTPUT_DIR / "hmrc_tdr_source_inventory.json")
    export_raw_rules(rules, OUTPUT_DIR / "hmrc_tdr_rules.json")
    export_master_pdf(dataset, OUTPUT_DIR / "hmrc_tdr_master_requirements_document.pdf")

    manifest = {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "output_dir": str(OUTPUT_DIR),
        "files": sorted(p.name for p in OUTPUT_DIR.iterdir() if p.is_file()),
        "rule_count": len(rules),
        "source_count": len(sources),
    }
    (OUTPUT_DIR / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    print(json.dumps(manifest, indent=2))


if __name__ == "__main__":
    main()
